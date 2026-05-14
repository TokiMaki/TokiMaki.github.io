#!/usr/bin/env python3

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


WORKBOOK_PATH = Path("Docs/OATH_VALUE_REVERSE_TABLE.xlsx")
BACKUP_PATH = Path("Docs/OATH_VALUE_REVERSE_TABLE.xlsx.openpyxl.bak")
SHEET_NAME = "서약 세트 단계"


def apply_formulas() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(WORKBOOK_PATH)
    if not BACKUP_PATH.exists():
        shutil.copy2(WORKBOOK_PATH, BACKUP_PATH)

    workbook = load_workbook(WORKBOOK_PATH)
    sheet = workbook[SHEET_NAME]

    headers = {
        "J1": "단계 증가량(%)",
        "K1": "누적 증가량(%)",
        "L1": "결정배율",
        "M1": "서약등급배율",
        "N1": "가호최종뎀배율",
        "O1": "서약스탯배율",
        "P1": "보정점수",
    }
    header_fill = PatternFill(fill_type="solid", fgColor="EAF2FF")
    for cell_ref, value in headers.items():
        cell = sheet[cell_ref]
        cell.value = value
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Q열 이후에 이전 시도에서 남은 값이 있으면 사용자 입력 영역과 혼동되지 않도록 비운다.
    for row in range(1, 23):
        for col in range(17, min(sheet.max_column, 20) + 1):
            sheet.cell(row=row, column=col).value = None

    for row in range(2, 23):
        if sheet[f"A{row}"].value in (None, "") or sheet[f"B{row}"].value in (None, ""):
            continue

        sheet[f"L{row}"] = (
            f"=POWER(1.13,N(E{row}))*POWER(1.17,N(F{row}))"
            f"*POWER(1.2,N(G{row}))*POWER(1.23,N(H{row}))*POWER(1.26,N(I{row}))"
        )
        sheet[f"M{row}"] = (
            f'=IF(D{row}="유니크",1.38,IF(D{row}="레전더리",1.44,IF(D{row}="에픽",1.5,IF(D{row}="태초",1.57,1))))'
        )
        sheet[f"N{row}"] = (
            f'=1+IF(C{row}>=2550,62+INT((C{row}-2550)/25)*0.5,'
            f'IF(C{row}>=2100,50+INT((C{row}-2100)/25)*0.5,'
            f'IF(C{row}>=1650,39+INT((C{row}-1650)/25)*0.5,'
            f'IF(C{row}>=1200,28+INT((C{row}-1200)/25)*0.5,'
            f'IF(C{row}>=750,17+INT((C{row}-750)/25)*0.5,0)))))/100'
        )
        sheet[f"O{row}"] = (
            f'=IF(D{row}="유니크",'
            f'(1+(6679+298+168350+297900+INT(3.08*(6679+298-815)+2886))/250)'
            f'/(1+(6679+168350+297900+INT(3.08*(6679-815)+2886))/250),'
            f'IF(D{row}="레전더리",'
            f'(1+(6679+350+168350+297900+INT(3.08*(6679+350-815)+2886))/250)'
            f'/(1+(6679+168350+297900+INT(3.08*(6679-815)+2886))/250),'
            f'IF(D{row}="에픽",'
            f'(1+(6679+400+168350+297900+INT(3.08*(6679+400-815)+2886))/250)'
            f'/(1+(6679+168350+297900+INT(3.08*(6679-815)+2886))/250),'
            f'IF(D{row}="태초",'
            f'(1+(6679+450+168350+297900+INT(3.08*(6679+450-815)+2886))/250)'
            f'/(1+(6679+168350+297900+INT(3.08*(6679-815)+2886))/250),1))))'
        )
        sheet[f"P{row}"] = f"=B{row}/L{row}/M{row}/N{row}/O{row}"
        sheet[f"J{row}"] = "=0" if row == 2 else f"=(P{row}/P{row - 1}-1)*100"
        sheet[f"K{row}"] = f"=(P{row}/$P$2-1)*100"

        for col in range(10, 16):
            sheet.cell(row=row, column=col).number_format = "0.0"
        sheet[f"P{row}"].number_format = "0.00"

    widths = {
        "J": 16,
        "K": 16,
        "L": 14,
        "M": 16,
        "N": 18,
        "O": 16,
        "P": 14,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width

    workbook.save(WORKBOOK_PATH)
    print(WORKBOOK_PATH)


if __name__ == "__main__":
    apply_formulas()
